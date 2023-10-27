import openpyxl

#Criar uma planilha
book = openpyxl.Workbook()

#Como visualizar páginas existentes
print(book.sheetnames)

#Como criar uma página
book.create_sheet("computadores")
#como selecionar uma página
computadores_page = book['computadores']
computadores_page.append(['ELETRÔNICA', 'MEMÓRIA RAM', 'PREÇO'])
computadores_page.append(['COMPUTADOR 1', '8GB RAM', 'R$2.500,00'])
computadores_page.append(['COMPUTADOR 2', '16GB RAM', 'R$5.550,00'])
computadores_page.append(['COMPUTADOR 3', '32GB RAM', 'R$8.500,00'])
#Salvar a planilha
book.save('Planilha de Computadores.xlsx')